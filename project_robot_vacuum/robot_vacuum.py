import openpyxl
import random
import sys
import time
from tkinter import messagebox
import os
import numpy as np


excel_file = openpyxl.load_workbook("excel_robot.xlsx") # Dosyanın konumu değişiklik gösterebilir
page = excel_file["Sayfa1"]

maxColumn = page.max_column
maxRow = page.max_row

sys.setrecursionlimit(200000)

# G2 - (2,7)

positionGarbageCollector = np.array([0,0])

# 0-Above 1-Right 2-Left 3-Down 
direction = 3 #Robotun başlangıçta hangi yöne doğru ilerlemesini istiyorsan.

positionRobot = np.array([2,7])
positionEyes = np.array([2,7])

#Gidilmemis Zemin:1 Robotun Konumu:2, Çöp:3, Temiz Zemin:4, Kırmızı Engel:9, Siyah Duvar:0

amountGarbage = 0
garbageCollected = 0
counter = 0


def lookAbove():
    global direction
    positionRobot[0] -= 1
    if (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 9) or (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 0):
        # print("There is an obstacle at point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
        f = open("moveCommend.txt","a")
        f.write("There is an obstacle at point {}".format(str(page.cell(row= positionRobot[0], column= positionRobot[1]))))
        f.write("\n")
        f.close()
        positionRobot[0] += 1
        target = random.randint(0,2)
        if target==0:
                direction = 3
                lookLeft()
                
        elif target==1:
                direction = 1
                lookRight()
                
        elif target==2:
                direction = 2
                lookDown()
    
    elif (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 3):
        cleanGarbage()
        positionRobot[0] += 1
        moveAbove()
    else:
        positionRobot[0] += 1
        moveAbove()

def lookDown():
    global direction
    positionRobot[0] += 1
    if (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 9) or (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 0):
        # print("There is an obstacle at point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
        f = open("moveCommend.txt","a")
        f.write("There is an obstacle at point {}".format(str(page.cell(row= positionRobot[0], column= positionRobot[1]))))
        f.write("\n")
        f.close()
        positionRobot[0] -= 1
        target = random.randint(0,2)
        if target==0:
                direction=3
                lookLeft()
                
        elif target==1:
                direction=1
                lookRight()
                
        elif target==2:
                direction=0
                lookAbove()        
        
    elif (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 3):
        cleanGarbage()
        positionRobot[0] -= 1
        moveDown()
    else:
        positionRobot[0] -= 1
        moveDown()
        
def lookRight():
    global direction
    positionRobot[1] += 1
    if (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 9) or (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 0):
        # print("There is an obstacle at point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
        f = open("moveCommend.txt","a")
        f.write("There is an obstacle at point {}".format(str(page.cell(row= positionRobot[0], column= positionRobot[1]))))
        f.write("\n")
        f.close()
        positionRobot[1] -= 1
        target = random.randint(0,2)
        if target==0:
                direction=2
                lookDown()
        
        elif target==1:
                direction=0
                lookAbove()
                
        elif target==2:
                direction=3
                lookLeft()

    elif (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 3):
        cleanGarbage()
        positionRobot[1] -= 1
        moveRight()
        
    else:
        positionRobot[1] -= 1
        moveRight()

def lookLeft():
    global direction
    positionRobot[1] -= 1
    if (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 9) or (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 0):
        # print("There is an obstacle at point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
        f = open("moveCommend.txt","a")
        f.write("There is an obstacle at point {}".format(str(page.cell(row= positionRobot[0], column= positionRobot[1]))))
        f.write("\n")
        f.close()
        positionRobot[1] += 1
        target = random.randint(0,2)
        if target==0:
                direction=2
                lookDown()
                
        elif target==1:
                direction=0
                lookAbove()
                
        elif target==2:
                direction=1
                lookRight()        

    elif (page.cell(row= positionRobot[0], column= positionRobot[1]).value == 3):
        cleanGarbage()
        positionRobot[1] += 1
        moveLeft()
        
    else:
        positionRobot[1] += 1
        moveLeft()

                
def directions():
    global direction
    if garbageCollected == amountGarbage:
        return 0
    
    if direction == 0:
         lookAbove()
    elif direction == 1:
         lookRight()
    elif direction ==2:
         lookDown()
    elif direction == 3 :
         lookLeft()
         
         
def Eyes_Above():
    if garbageCollected == amountGarbage:
        return 0
    
    eyesAbove = True
    global positionEyes
    while eyesAbove == True:
        positionEyes[0]-=1
        # print("Eyes moved upward to point {}".format(page.cell(row= positionEyes[0], column= positionEyes[1])))
        if(page.cell(row=positionEyes[0],column=positionEyes[1]).value==3):
            # print(" *^* ")
            lookAbove()
        elif (page.cell(row=positionEyes[0],column=positionEyes[1]).value==0) or (page.cell(row=positionEyes[0],column=positionEyes[1]).value==9):
            positionEyes[0]=positionRobot[0]
            # print(" ^ ")
            eyesAbove = False
        Eyes_Right()
          
def Eyes_Down():
    if garbageCollected == amountGarbage:
        return 0
    eyesDown = True
    global positionEyes
    while eyesDown == True:
        positionEyes[0]+=1
        # print("Eyes moved down to point {}".format(page.cell(row= positionEyes[0], column= positionEyes[1])))
        if (page.cell(row= positionEyes[0],column= positionEyes[1]).value==3):
            # print(" *v*")
            lookDown()
        elif (page.cell(row=positionEyes[0],column=positionEyes[1]).value==0) or (page.cell(row=positionEyes[0],column=positionEyes[1]).value==9):
            positionEyes[0]=positionRobot[0]
            # print(" v ")
            eyesDown = False
        Eyes_Left()
                        
def Eyes_Left():
    if garbageCollected == amountGarbage:
        return 0
    eyesLeft = True
    global positionEyes
    while eyesLeft == True:
        positionEyes[1]-=1
        # print("Eyes moved to the left to point {}".format(page.cell(row= positionEyes[0], column= positionEyes[1])))
        if(page.cell(row=positionEyes[0],column=positionEyes[1]).value==3):
            # print(" *<*")
            lookLeft()
        elif (page.cell(row=positionEyes[0],column=positionEyes[1]).value==0) or (page.cell(row=positionEyes[0],column=positionEyes[1]).value==9):
            positionEyes[1]=positionRobot[1]
            # print(" < ")
            eyesLeft = False
        directions()
            
def Eyes_Right():
    if garbageCollected == amountGarbage:
        return 0
    eyesRight = True
    global positionEyes
    while eyesRight == True:
        positionEyes[1]+=1
        # print("Eyes moved to the right to point {}".format(page.cell(row= positionEyes[0], column= positionEyes[1])))
        if(page.cell(row=positionEyes[0],column=positionEyes[1]).value==3):
            # print(" *>*")
            lookRight()
        elif (page.cell(row=positionEyes[0],column=positionEyes[1]).value==0) or (page.cell(row=positionEyes[0],column=positionEyes[1]).value==9):
            positionEyes[1]=positionRobot[1]
            # print(" > ")
            eyesRight = False
        Eyes_Down()
            
                     
def moveAbove():
    time.sleep(0.01)
    page.cell(row= positionRobot[0], column= positionRobot[1], value="4")
    positionRobot[0] -= 1
    positionEyes[0] -= 1
    page.cell(row= positionRobot[0], column= positionRobot[1], value="2")
    # print("moved upwards to point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
    writeCommend()
    global counter
    counter += 1
    print(garbageCollected)
    print(amountGarbage)
    if garbageCollected == amountGarbage:
        return 0 
    Eyes_Above()

def moveDown():
    time.sleep(0.01)
    page.cell(row= positionRobot[0], column= positionRobot[1], value="4")
    positionRobot[0] += 1
    positionEyes[0] += 1
    page.cell(row= positionRobot[0], column= positionRobot[1], value="2")
    # print("moved down to point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
    writeCommend()
    global counter
    counter += 1
    print(garbageCollected)
    print(amountGarbage)
    if garbageCollected == amountGarbage:
        return 0 
    Eyes_Above()

def moveRight():
    time.sleep(0.01)
    page.cell(row= positionRobot[0], column= positionRobot[1], value="4")
    positionRobot[1] += 1
    positionEyes[1] += 1
    page.cell(row= positionRobot[0], column= positionRobot[1], value="2")
    # print("moved to the right to point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
    writeCommend()
    global counter
    counter += 1
    print(garbageCollected)
    print(amountGarbage)
    if garbageCollected == amountGarbage:
        return 0 
    Eyes_Above()

def moveLeft():
    time.sleep(0.01)
    page.cell(row= positionRobot[0], column= positionRobot[1], value="4")
    positionRobot[1] -= 1
    positionEyes[1] -= 1
    page.cell(row= positionRobot[0], column= positionRobot[1], value="2")
    # print("moved to the left to point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
    writeCommend()
    global counter
    counter += 1
    print(garbageCollected)
    print(amountGarbage)
    if garbageCollected == amountGarbage:
        return 0 
    Eyes_Above()


def writeCommend():
    f = open("moveCommend.txt","a")
    f.write("moved to the point {}".format(str(page.cell(row= positionRobot[0], column= positionRobot[1]))))
    f.write("\n")
    f.close()

def cleanGarbage():
    global garbageCollected
    # print("Garbage cleared at point {}".format(page.cell(row= positionRobot[0], column= positionRobot[1])))
    g = open("garbageCommend.txt","a")
    g.write("Garbage cleared at point {}".format(str(page.cell(row= positionRobot[0], column= positionRobot[1]))))
    g.write("\n")
    g.close()
    time.sleep(1)
    garbageCollected += 1

def garbageCounter():
    global positionGarbageCollector
    global amountGarbage
    x=0
    y=0
    while x < maxRow:
        x += 1
        positionGarbageCollector[0]+=1
        y = 0
        positionGarbageCollector[1]=0
        while y < maxColumn:
            y += 1
            positionGarbageCollector[1]+=1
            if (page.cell(row= positionGarbageCollector[0], column= positionGarbageCollector[1]).value == 3):
                amountGarbage +=1


def main():
    os.remove("moveCommend.txt")
    os.remove("garbageCommend.txt")
    lookDown()
    if garbageCollected == amountGarbage:
        return 0 

garbageCounter()   
main()

messagebox.showinfo(
    "Cleaning Complete","{} Garbage Cleaned Up in {} Steps".format(amountGarbage,counter)
)